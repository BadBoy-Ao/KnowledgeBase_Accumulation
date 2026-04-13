import os
import uuid
import zipfile
import json
from openpyxl import Workbook
import xml.etree.ElementTree as ET
from io import BytesIO
import shutil

'''
# 在 HTTP.py 开头添加
import sys
from pathlib import Path

# 获取当前文件的父目录（即 HTTP 目录）的父目录（即 app 目录）
parent_dir = str(Path(__file__).parent.parent)
if parent_dir not in sys.path:
    print(f"{parent_dir}")
    sys.path.append(parent_dir)
'''

# def create_cellimages_xml(disp_id, rId):
def create_cellimages_xml(updated_pics_data):
    # 定义命名空间 URI
    NS_ETC = "http://www.wps.cn/officeDocument/2017/etCustomData"
    NS_XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    NS_A = "http://schemas.openxmlformats.org/drawingml/2006/main"
    NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

    # 注册命名空间前缀
    ET.register_namespace("etc", NS_ETC)
    ET.register_namespace("xdr", NS_XDR)
    ET.register_namespace("a", NS_A)
    ET.register_namespace("r", NS_R)

    # 创建根元素
    root = ET.Element(f"{{{NS_ETC}}}cellImages")

    # 子元素：全部使用 {uri} 格式
    for i, pic_info in enumerate(updated_pics_data):
        cell_image = ET.SubElement(root, f"{{{NS_ETC}}}cellImage")
        pic = ET.SubElement(cell_image, f"{{{NS_XDR}}}pic")

        nvPr = ET.SubElement(pic, f"{{{NS_XDR}}}nvPicPr")
        cNvPr = ET.SubElement(nvPr, f"{{{NS_XDR}}}cNvPr")
        cNvPr.set("id", f"{i + 1}")
        cNvPr.set("name", pic_info["disp_id"])
        ET.SubElement(nvPr, f"{{{NS_XDR}}}cNvPicPr")

        blipFill = ET.SubElement(pic, f"{{{NS_XDR}}}blipFill")
        blip = ET.SubElement(blipFill, f"{{{NS_A}}}blip")
        blip.set(f"{{{NS_R}}}embed", f"rId{i + 1}")
        stretch = ET.SubElement(blipFill, f"{{{NS_A}}}stretch")
        ET.SubElement(stretch, f"{{{NS_A}}}fillRect")

        spPr = ET.SubElement(pic, f"{{{NS_XDR}}}spPr")
        xfrm = ET.SubElement(spPr, f"{{{NS_A}}}xfrm")
        off_a = ET.SubElement(xfrm, f"{{{NS_A}}}off")
        off_a.set("x", "0")
        off_a.set("y", "0")
        ext_a = ET.SubElement(xfrm, f"{{{NS_A}}}ext")
        ext_a.set("cx", "5000000") # 高度
        ext_a.set("cy", "3000000") # 宽度
        prstGeom = ET.SubElement(spPr, f"{{{NS_A}}}prstGeom")
        prstGeom.set("prst", "rect")
        ET.SubElement(prstGeom, f"{{{NS_A}}}avLst")

    # 输出 XML 字符串
    xml_str = ET.tostring(root, encoding='unicode')
    return '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n' + xml_str


# def create_cellimages_xml_rels(disp_img, rId):
def create_cellimages_xml_rels(updated_pics_data):
    NS = "http://schemas.openxmlformats.org/package/2006/relationships"
    ET.register_namespace('', NS)

    root = ET.Element("Relationships", xmlns=NS)
    for i, disp_info in enumerate(updated_pics_data):
        rel = ET.SubElement(root, "Relationship")
        rel.set("Id", f"rId{i + 1}")
        rel.set("Type", "http://schemas.openxmlformats.org/officeDocument/2006/relationships/image")
        rel.set("Target", disp_info["image_zip_path"][3:])

    return '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>' + ET.tostring(root, encoding='unicode')


def write_pics_data_to_excel(sheet, pics_data, json_data, current_dir, start_col=1):
    """
    将图片数据和JSON数据写入Excel工作表
    第一行：图片类型 + JSON键
    第二行：图片公式 + JSON值
    同时将生成的UUID存储到pics_data中
    """
    col = start_col

    # 写入JSON数据
    for key, value in json_data.items():
        sheet.cell(row=1, column=col, value=key)  # 第一行写入键
        sheet.cell(row=2, column=col, value=value)  # 第二行写入值
        col += 1

    # 写入图片数据
    for i, pic_info in enumerate(pics_data):
        pic_type = pic_info["图片类型"]
        sheet.cell(row=1, column=col, value=pic_type)  # 第一行写入图片类型

        # 生成唯一的图片ID并写入公式
        disp_id = f"ID_{uuid.uuid4().hex.upper()}"
        sheet.cell(row=2, column=col, value=f'=DISPIMG("{disp_id}",1)')  # 第二行写入公式

        # 将生成的UUID存储到pics_data中
        pic_info["disp_id"] = disp_id

        # 修改文件路径
        pic_info["图片文件路径"] = os.path.join(current_dir, pic_info["图片文件路径"])
        col += 1

    return pics_data, col  # 返回更新后的图片信息和下一列的列号


def JSON2EXCEL_Function(pics_data, json_data, output_xlsx):
    sheet_name = "sheet1"

    # 获取当前文件所在目录（即函数定义的目录）
    current_dir_0 = os.path.dirname(os.path.realpath(__file__))
    current_dir = os.path.join(current_dir_0, "received_files")
    # print("函数所在目录:", current_dir)

    # 检查所有图片文件是否存在
    for pic_info in pics_data:
        pic_info_path = os.path.join(current_dir, pic_info["图片文件路径"])
        if not os.path.exists(pic_info_path):
            print(f"❌ 错误：图片路径不存在 {pic_info_path}")
            return

    # 创建工作簿并写入数据
    wb = Workbook()
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
    else:
        sheet = wb.create_sheet(sheet_name)
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

    # 写入图片数据和JSON数据到Excel，并存储UUID
    updated_pics_data, next_col = write_pics_data_to_excel(sheet, pics_data, json_data, current_dir, start_col=1)
    print(f"✅ 图片数据和JSON已写入Excel，UUID已存储到pics_data中")

    # 打印生成的UUID
    for i, pic_info in enumerate(updated_pics_data):
        print(f"  图片 {i + 1}: {pic_info['图片类型']} -> UUID: {pic_info['disp_id']}")

    # 保存为临时文件
    temp_file = "temp.xlsx"
    wb.save(temp_file)

    # 重建ZIP文件，添加图片功能
    new_temp = "new_temp.xlsx"
    with (zipfile.ZipFile(temp_file, 'r') as old_zf):
        with zipfile.ZipFile(new_temp, 'w', zipfile.ZIP_DEFLATED) as new_zf:
            # 复制所有原有文件，跳过我们要修改的
            for item in old_zf.infolist():
                if item.filename in {
                    "xl/_rels/workbook.xml.rels",
                    "[Content_Types].xml",
                    "xl/cellimages.xml",
                    "xl/_rels/cellimages.xml.rels"
                }:
                    continue  # 稍后重写
                new_zf.writestr(item, old_zf.read(item))

            # 添加所有图片
            image_files = []
            for i, pic_info in enumerate(updated_pics_data):
                pic_info_url = pic_info["图片文件路径"]
                # 读取图片内容
                with open(f"{pic_info_url}", "rb") as f:
                    image_bytes = f.read()

                # 获取图片扩展名
                filename = os.path.basename(pic_info["图片文件路径"]).lower()
                if filename.endswith(".png"):
                    ext = "png"
                elif filename.endswith((".jpg", ".jpeg")):
                    ext = "jpeg"
                else:
                    ext = "jpeg"

                # 添加图片到ZIP
                image_zip_path = f"xl/media/image{i + 1}.{ext}"
                new_zf.writestr(image_zip_path, image_bytes)
                image_files.append((image_zip_path, ext, pic_info["disp_id"]))  # 添加disp_id信息
                # 将属性写入列表对象
                pic_info["image_zip_path"] = image_zip_path
                pic_info["ext"] = ext

            # 为第一个图片创建cellimages.xml（简化处理）
            # 实际应用中可能需要根据实际情况创建更复杂的结构
            # disp_id = updated_pics_data[0]["disp_id"]  # 使用第一个图片的UUID
            # cellimages_xml = create_cellimages_xml(disp_id, "rId1")
            '''导入对象列表，进行循环'''
            cellimages_xml = create_cellimages_xml(updated_pics_data)
            new_zf.writestr("xl/cellimages.xml", cellimages_xml)

            '''
            # 添加 cellimages.xml.rels
            '''
            # cellimages_rels = create_cellimages_xml_rels("media/image1.png", "rId1")
            cellimages_rels = create_cellimages_xml_rels(updated_pics_data)
            new_zf.writestr("xl/_rels/cellimages.xml.rels", cellimages_rels)


            # 构建 [Content_Types].xml
            NS_CT = "http://schemas.openxmlformats.org/package/2006/content-types"
            ET.register_namespace('', NS_CT)

            ct_root = ET.Element("Types", xmlns=NS_CT)

            # 复制旧的（如果存在）
            try:
                ct_old = old_zf.read("[Content_Types].xml").decode('utf-8')
                ct_root = ET.fromstring(ct_old)
            except Exception:
                pass  # 从空开始

            # 确保 cellimages.xml 类型存在
            has_cellimages = False
            for override in ct_root.findall(f"{{{NS_CT}}}Override"):
                if override.get("PartName") == "/xl/cellimages.xml":
                    has_cellimages = True
                    break
            if not has_cellimages:
                override = ET.SubElement(ct_root, f"{{{NS_CT}}}Override")
                override.set("PartName", "/xl/cellimages.xml")
                override.set("ContentType", "application/vnd.wps-officedocument.cellimage+xml")

            # 确保所有图片类型存在
            for image_zip_path, ext, _ in image_files:
                media_part = f"/xl/{image_zip_path[3:]}"  # 去掉xl/前缀
                has_image = False
                for override in ct_root.findall(f"{{{NS_CT}}}Override"):
                    if override.get("PartName") == media_part:
                        has_image = True
                        break
                if not has_image:
                    override = ET.SubElement(ct_root, f"{{{NS_CT}}}Override")
                    override.set("PartName", media_part)
                    override.set("ContentType", f"image/{'png' if ext == 'png' else 'jpeg'}")

            ct_xml = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n' + ET.tostring(ct_root, encoding='unicode')
            new_zf.writestr("[Content_Types].xml", ct_xml)

            # 构建 workbook.xml.rels
            rels_path = "xl/_rels/workbook.xml.rels"
            ns = "http://schemas.openxmlformats.org/package/2006/relationships"
            ET.register_namespace('', ns)

            try:
                rels_data = old_zf.read(rels_path).decode('utf-8')
                root = ET.fromstring(rels_data)
            except KeyError:
                root = ET.Element("Relationships", xmlns=ns)

            # 检查是否已存在 cellimages.xml 关系
            cellimages_exists = False
            max_rid = 0
            for rel in root.findall(f"{{{ns}}}Relationship"):
                target = rel.get("Target", "")
                rid = rel.get("Id", "")
                if target == "cellimages.xml":
                    cellimages_exists = True
                if rid.startswith("rId") and rid[3:].isdigit():
                    num = int(rid[3:])
                    max_rid = max(max_rid, num)

            if not cellimages_exists:
                new_id = f"rId{max_rid + 1}"
                rel = ET.SubElement(root, f"{{{ns}}}Relationship")
                rel.set("Id", new_id)
                rel.set("Type", "http://www.wps.cn/officeDocument/2020/cellImage")
                rel.set("Target", "cellimages.xml")

            rels_xml = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>' + ET.tostring(root, encoding='unicode')
            new_zf.writestr(rels_path, rels_xml)

    # 替换为最终文件
    '''
    os.replace() 和 os.rename() 在底层都依赖于操作系统的 rename() 系统调用，
    这个调用要求源文件和目标文件必须在同一个文件系统上才能进行原子性的“移动/替换”操作。
    '''
    # os.replace(new_temp, output_xlsx)

    shutil.move(new_temp, output_xlsx)
    os.remove(temp_file)  # 清理临时文件

    print(f"✅ 生成完成：{output_xlsx}")
    print("第一行: JSON键 + 图片类型")
    print("第二行: JSON值 + 图片公式")
    print("请用 WPS 打开查看")
    return output_xlsx


if __name__ == "__main__":
    # received_files/
    pics_data = [
        {
            "图片名称": "测试集_图片_社会团体法人登记证书】数据_图片_01.png",
            "图片文件路径": "测试集_图片_社会团体法人登记证书】数据_图片_01.png",
            "图片类型": "营业执照"
        },
        {
            "图片名称": "测试集_图片_社会团体法人登记证书】数据_图片_01.png",
            "图片文件路径": "测试集_图片_社会团体法人登记证书】数据_图片_01.png",
            "图片类型": "身份证正面"
        }
    ]

    json_data = {
        "签名1": "敖小胖",
        "签名2": "敖小胖",
        "签名3": "敖小胖",
        "签名4": "敖小胖"
    }

    output_xlsx = "t2.xlsx"
    JSON2EXCEL_Function(pics_data,json_data,output_xlsx)


